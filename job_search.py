"""
================================================================================
  JOB SEARCH  —  find fresh job postings and save them to a clean Excel file
================================================================================

  HOW TO USE:
    1. Edit the SETTINGS section below (job titles + location).
    2. Run it:   python job_search.py
    3. Open the Excel file it creates in this folder.

  You do NOT need to touch any code below the SETTINGS section.
================================================================================

  Powered by JobSpy (https://github.com/speedyapply/JobSpy), MIT-licensed.
  JobSpy does the job-board searching; this script wraps it in a simple config
  and a clean, formatted Excel output.
================================================================================
"""

# ==============================================================================
#  SETTINGS  —  this is the only part you need to edit
# ==============================================================================

# 1) The job titles you want to search for. Add or remove lines freely.
#    Put each title in quotes, with a comma at the end of the line.
JOB_TITLES = [
    "Sales Representative",
    "Medical Sales Representative",
    "Account Executive",
    "Business Development Representative",
]

# 2) Where you want to search. Examples:
#      "San Francisco, CA"   "New York, NY"   "Remote"   "Chicago, IL"
LOCATION = "San Francisco, CA"

# 3) How far back to look, in HOURS. (Tip: number of days x 24)
#      24 = last day  |  72 = last 3 days  |  168 = last week  |  336 = last 2 weeks
HOURS_OLD = 72

# 4) How many results to pull per site, per title. 50 is a good default.
#    Higher = more results but slower and more likely to be rate-limited.
RESULTS_PER_SITE = 50

# 5) (Optional) Skip any posting whose title contains one of these words.
#    Leave the list empty  ->  []  to keep everything.
#    Example: ["Senior", "Manager", "Director"]
EXCLUDE_TITLE_KEYWORDS = []

# 6) Which job boards to search. Remove any you don't want.
SITES = ["indeed", "linkedin", "zip_recruiter", "glassdoor"]

# 7) Country (only used by Indeed & Glassdoor). Usually leave as "USA".
COUNTRY = "USA"

# ==============================================================================
#  You can stop reading here. Everything below just makes it work.
# ==============================================================================

import sys
from datetime import datetime

# --- Friendly check that the required packages are installed -------------------
try:
    import pandas as pd
    from jobspy import scrape_jobs
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\n  It looks like the required packages aren't installed yet.")
    print("  Open a terminal in this folder and run:\n")
    print("      pip install -r requirements.txt\n")
    sys.exit(1)


# The columns we keep, in the order they'll appear in the spreadsheet.
# (left = raw column from the data source, right = the friendly header you'll see)
OUTPUT_COLUMNS = {
    "title": "Job Title",
    "company": "Company",
    "location": "Location",
    "Salary": "Salary",
    "job_type": "Job Type",
    "Remote": "Remote?",
    "date_posted": "Date Posted",
    "site": "Source",
    "job_url": "Link",
    "description": "Description",
}


def format_salary(row):
    """Turn the raw min/max/interval columns into one readable string."""
    low = row.get("min_amount")
    high = row.get("max_amount")
    interval = row.get("interval")
    interval = "" if pd.isna(interval) else str(interval).lower()

    if pd.isna(low) and pd.isna(high):
        return "Not listed"

    suffix = {
        "yearly": "/yr",
        "hourly": "/hr",
        "monthly": "/mo",
        "weekly": "/wk",
        "daily": "/day",
    }.get(interval, "")

    def money(v):
        return f"${int(v):,}" if pd.notna(v) else ""

    if pd.notna(low) and pd.notna(high):
        return f"{money(low)} – {money(high)}{suffix}"
    return f"{money(low or high)}{suffix}"


def scrape_all_titles():
    """Search every job title and return one combined DataFrame."""
    collected = []

    for title in JOB_TITLES:
        print(f"  Searching: {title} ...", end=" ", flush=True)
        try:
            results = scrape_jobs(
                site_name=SITES,
                search_term=title,
                location=LOCATION,
                results_wanted=RESULTS_PER_SITE,
                hours_old=HOURS_OLD,
                country_indeed=COUNTRY,
            )
        except Exception as e:
            print(f"skipped (error: {e})")
            continue

        if results is not None and len(results) > 0:
            collected.append(results)
            print(f"found {len(results)}")
        else:
            print("found 0")

    if not collected:
        return pd.DataFrame()

    return pd.concat(collected, ignore_index=True)


def clean_results(df):
    """Remove duplicates, apply filters, and build the final tidy columns."""
    if df.empty:
        return df

    # Remove exact duplicate postings (same link, or same title+company+location).
    if "job_url" in df.columns:
        df = df.drop_duplicates(subset=["job_url"])
    df = df.drop_duplicates(subset=["title", "company", "location"])

    # Optional: drop titles containing any excluded keyword.
    for word in EXCLUDE_TITLE_KEYWORDS:
        df = df[~df["title"].str.contains(word, case=False, na=False)]

    # Build the derived columns.
    df["Salary"] = df.apply(format_salary, axis=1)
    if "is_remote" in df.columns:
        df["Remote"] = df["is_remote"].map({True: "Yes", False: "No"}).fillna("")
    else:
        df["Remote"] = ""

    # Keep only the columns we want, in order (skip any the source didn't provide).
    keep = [c for c in OUTPUT_COLUMNS if c in df.columns]
    df = df[keep].rename(columns=OUTPUT_COLUMNS)

    # Some sites (e.g. LinkedIn) don't include the description in the listing.
    # For those, point the reader to the posting instead of leaving a blank cell.
    if "Description" in df.columns:
        df["Description"] = df["Description"].fillna("").astype(str)
        missing = df["Description"].str.strip() == ""
        df.loc[missing, "Description"] = "Click link to find description"

    # Newest postings first.
    if "Date Posted" in df.columns:
        df = df.sort_values("Date Posted", ascending=False, na_position="last")

    return df.reset_index(drop=True)


def save_to_excel(df, path):
    """Write the DataFrame to a nicely formatted, easy-to-read Excel file."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Jobs")
        ws = writer.sheets["Jobs"]

        # --- Style the header row ---
        header_fill = PatternFill("solid", fgColor="1F4E78")   # dark blue
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[1].height = 22

        # Freeze the header row and turn on filter dropdowns.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # --- Set sensible column widths ---
        widths = {
            "Job Title": 38, "Company": 26, "Location": 24, "Salary": 20,
            "Job Type": 14, "Remote?": 9, "Date Posted": 13, "Source": 13,
            "Link": 55, "Description": 70,
        }
        headers = [c.value for c in ws[1]]
        for i, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 18)

        # --- Make the Link column clickable ---
        if "Link" in headers:
            link_col = headers.index("Link") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=link_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")


def main():
    print("\n" + "=" * 60)
    print("  JOB SEARCH")
    print(f"  Location: {LOCATION}   |   Last {HOURS_OLD} hours")
    print("=" * 60 + "\n")

    raw = scrape_all_titles()

    if raw.empty:
        print("\n  No jobs found. Try widening HOURS_OLD or changing the titles/location.")
        print("  (The job boards may also be rate-limiting — wait a bit and retry.)\n")
        return

    df = clean_results(raw)

    if df.empty:
        print("\n  All results were filtered out. Check EXCLUDE_TITLE_KEYWORDS.\n")
        return

    # Save with a date-stamped name so you never overwrite an earlier run.
    filename = f"Job_Postings_{datetime.now():%Y-%m-%d_%H%M}.xlsx"
    save_to_excel(df, filename)

    print("\n" + "=" * 60)
    print(f"  Done! {len(df)} unique job postings saved to:")
    print(f"     {filename}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
